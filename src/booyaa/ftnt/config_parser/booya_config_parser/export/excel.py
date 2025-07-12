from datetime import datetime
from openpyxl import Workbook, styles, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from os import makedirs, path
from pprint import pprint
from pathlib import Path
from xml.etree import ElementTree as ET
from booya_config_parser.export import horizontal, virtical, ha
from booya_config_parser.export.style import write_sheet_title
from export.excel_font_change import base_font_change

class ExportExcel():
    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb.worksheets[0])

    def timestamp(self):
        return datetime.now(datetime.now().astimezone().tzinfo).strftime('%Y%m%d_%H%M%S')

    def gen_book_data(self, formatted_config_obj):
        """
        {
            'version': version,
            'model': model,
            'major': major,
            'minor': minor,
            'patch': patch,
            'conf_file_ver': self.config_obj['header']['conf_file_ver'],
            'opmode': self.config_obj['header']['opmode'],
            'operation_mode': operation_mode,
            'user': self.config_obj['header']['user'],
            'vdom_mode': vdom_mode,
            'vdom': self.config_obj['header']['vdom'],
            'global_vdom': self.config_obj['header']['global_vdom'],
            'hostname': self.config_obj[global_key]['system_global']['hostname'],
        }
        """
        header = formatted_config_obj['header']
        self.version = header['version']
        self.hostname = header['hostname']
        self.major = header['major']
        self.minor = header['minor']
        self.patch = header['patch']
        self.vdom_mode = header['vdom_mode']
        self.book_name = f'{self.hostname}_{self.version}_{self.timestamp()}'

    def export_book(self, formatted_config_obj, vdom_field_names, global_field_names, export_dir='./tmp/export'):
        """"""
        makedirs(export_dir, exist_ok=True)
        self.export_dir = export_dir


        for vdom_name in formatted_config_obj:
            if vdom_name == 'header':
                self.gen_book_data(formatted_config_obj)
                continue

            if vdom_name == 'global':
                field_names = global_field_names
            else:
                field_names = vdom_field_names

            for category in field_names:
                # シート名
                # 31文字オーバーの対処 @todo
                if vdom_name == 'global':
                    sheetname = f"{field_names[category]['__layout__']['sheetname']}"
                else:
                    sheetname = f"{vdom_name} {field_names[category]['__layout__']['sheetname']}"

                # シート作成
                ws, row = self.add_sheet(sheetname)

                # 次の行に移動
                row = row + 1

                # ha用レイアウト
                if category == 'system_ha':
                    ha.write_table(
                        ws, row, vdom_name, category,
                        formatted_config_obj,
                        field_names)

                # ipsec vpn用レイアウト
                elif category == 'ipsec_vpn':
                    continue

                # address groupなど縦型汎用レイアウト
                elif field_names[category]['__layout__']['layout'] == 'virtical':
                    virtical.write_table(
                        ws, row, vdom_name, category,
                        formatted_config_obj[vdom_name][category],
                        field_names[category])

                # 汎用的な水平テーブルレイアウト
                else:
                    row += 1
                    # self.simple_table(ws, row, vdom_name, category, formatted_config_obj, field_names)
                    ws = horizontal.write_table(
                        ws, row, vdom_name, category,
                        formatted_config_obj[vdom_name][category],
                        field_names[category])

        self.save(font_change=True, font='Yu Gothic UI')

    def save(self, font_change=True, font='Yu Gothic UI'):
        book_path = Path(self.export_dir, f'{self.book_name}.xlsx')
        self.wb.save(book_path)

        if font_change:
            base_font_change(book_path, font)

    def add_sheet(self, sheetname, row=1, col=1):
        self.wb.create_sheet(sheetname)
        ws = self.wb[sheetname]
        # シートタイトル
        ws = write_sheet_title(ws, row, col, sheetname)

        return ws, row


if __name__ == '__main__':
    pass
