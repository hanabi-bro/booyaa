from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# excel style propety
table_header_fill = PatternFill(patternType='solid', fgColor='d3d3d3')
table_header_font = Font(bold=True)
# fill = PatternFill(patternType='lightHorizontal', fgColor='d3d3d3', bgColor='ff1493')
table_side = Side(style='thin', color='000000')
table_border = Border(
    top=table_side, bottom=table_side,
    left=table_side, right=table_side,
)
cell_aliment = Alignment(wrapText=True)

# sheet title font
sheet_title_font = Font(bold=True, size='16')


def write_sheet_title(ws, row, col, value):
    ws.cell(row=row, column=col).value = value
    ws.cell(row=row, column=col).font = sheet_title_font

    return ws


def write_hedder_cell(ws, row, col, value, width):
    ws.cell(row=row, column=col).value = value
    ws.cell(row=row, column=col).font = table_header_font
    ws.column_dimensions[get_column_letter(col)].width = width
    ws.cell(row=row, column=col).border = table_border
    ws.cell(row=row, column=col).fill = table_header_fill

    return ws


def write_cell(ws, row, col, value, width=None):
    ws.cell(row=row, column=col).value = value
    ws.cell(row=row, column=col).alignment = cell_aliment
    ws.cell(row=row, column=col).border = table_border
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

    return ws
