from datetime import datetime
from openpyxl import Workbook, workbook, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from pathlib import Path
from csv import DictReader
from config_parser.export.excel_style import write_sheet_title, write_hedder_cell, write_cell
from pprint import pprint


# bookオブジェクト作成
# フォーマット取得
# タイムスタンプ取得
# ブック名作成
# base sheet追加(header)
# タイトル追加
# データ書き込み
# global sheet追加
# タイトル追加
# データ書き込み
# vdom sheet追加
# タイトル追加
# データ書き込み




class ExportExcel:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb.worksheets[0])
        self.file_dir = Path(__file__).parent
        self.data_dir = Path(self.file_dir.parent, 'data')

    def timestamp(self) -> str:
        return datetime.now(datetime.now().astimezone().tzinfo).strftime('%Y%m%d_%H%M%S')

    def gen_book_data(self) -> None:
        header = self.formatted_config_obj['header']
        self.book_name = f'{header['hostname']}_{header['version']}_{self.timestamp()}.xlsx'

        self.load_format_data()

    def load_format_data(self) -> None:
        format_rule_global_csv = Path(self.rule_dir, r'format_rule_global.csv')
        with open(format_rule_global_csv, 'r', encoding='utf-8') as f:
            self.format_rule_global = list(DictReader(f))

        format_rule_vdom_csv = Path(self.rule_dir,  r'format_rule_vdom.csv')
        with open(format_rule_vdom_csv, 'r', encoding='utf-8') as f:
            self.format_rule_vdom = list(DictReader(f))

    def save(self, font_change=True, font='Yu Gothic UI') -> None:
        if not self.export_dir:
            Path.mkdir(self.export_dir, exist_ok=True)
        book_path = Path(self.export_dir, self.book_name)
        self.wb.save(book_path)

        # if font_change:
        #     base_font_change(book_path, font)

    def add_sheet_title(self, sheetname, title, row=1, col=1) ->  any:
        self.wb.create_sheet(sheetname)
        ws = self.wb[sheetname]
        ws.cell(row=row, column=col).value = title
        # ws.cell(row=row, column=col).font = sheet_title_font

        return ws

    def gen_category_sheet(self, rule_obj, global_confi_obj, mode, row=3, col=1) -> None:
        config_obj = global_confi_obj[rule_obj['category']]
        ws = self.add_sheet_title(rule_obj['sheet_name'], rule_obj['sheet_title'])
        if rule_obj['layout'] == 'vertical':
            self.write_vertical(ws, rule_obj, config_obj, row, col)
        elif rule_obj['layout'] == 'horizontal':
            pass
        else:
            print(f'unknown layout: {rule_obj['layout']}')

    def gen_vdom_category_sheet(self, rule_obj, vdom_obj, vdom_name='root', row=3, col=1) -> None:
        config_obj = vdom_obj[rule_obj['category']]
        ws = self.add_sheet_title(f'{vdom_name} {rule_obj['sheet_name']}', f'{vdom_name} {rule_obj['sheet_title']}')
        if rule_obj['layout'] == 'vertical':
            self.write_vertical(ws, rule_obj, config_obj, row, col)
        elif rule_obj['layout'] == 'horizontal':
            pass
        else:
            print(f'unknown layout: {rule_obj['layout']}')

    def write_vertical(self, ws, rule_obj, config_obj_list, row=3, col=1) -> None:
        # ws = self.add_sheet_title(rule_obj['sheet_name'], rule_obj['sheet_title'])

        # ルールを読み込み
        if rule_obj['type'] == 'global':
            format_file = Path(self.global_rule_dir, f'{rule_obj['category']}.csv')
        elif rule_obj['type'] == 'vdom':
            format_file = Path(self.vdom_rule_dir, f'{rule_obj['category']}.csv')
        with open(format_file, 'r', encoding='utf-8') as f:
            fmt = list(DictReader(f))

        # table title
        # 未実装

        # table header
        hedaer_row = row
        view_rule = [r for r in fmt if r['visible'].lower() != 'false']
        for i, r in enumerate(view_rule, start=1):
            if r['view_name']:
                value = r['view_name']
            else:
                value = r['view_key']

            write_hedder_cell(ws, hedaer_row, col, value, r['width'])
            hedaer_row = hedaer_row + 1

        # table data
        data_row = row
        col = col + 1
        if type(config_obj_list) is dict:
            config_obj_list = [config_obj_list]

        if rule_obj['plural'] == 'single':
            for data_col, config_obj in enumerate(config_obj_list, start=col):
                for j, r in enumerate(view_rule, start=1):
                    value = config_obj[r['config_name']]
                    if type(value) is list:
                        value = '\n'.join(config_obj[r['config_name']])

                    write_cell(ws, data_row, data_col, value)
                    data_row = data_row + 1

        elif rule_obj['plural'] == 'multi':
            for data_col, config_obj in enumerate(config_obj_list, start=col):
                data_row = row
                for j, r in enumerate(view_rule, start=1):
                    k = next(iter(config_obj.keys()))
                    value = config_obj[k][r['config_name']]
                    if type(value) is list:
                        value = '\n'.join(config_obj[k][r['config_name']])

                    write_cell(ws, data_row, data_col, value)
                    data_row = data_row + 1
        else:
            print(f'unknow plural type {rule_obj['plural']}')

    def write_horizontal(self, ws, rule_obj, config_obj, row=3, col=1) -> None:
        # 　カテゴリルールを読み込み
        format_file = Path(self.vdom_rule_dir, f'{rule_obj['category']}.csv')
        with open(format_file, 'r', encoding='utf-8') as f:
            fmt = list(DictReader(f))

        # ルールを読み込み
        if rule_obj['type'] == 'global':
            format_file = Path(self.global_rule_dir, f'{rule_obj['category']}.csv')
        elif rule_obj['type'] == 'vdom':
            format_file = Path(self.vdom_rule_dir, f'{rule_obj['category']}.csv')
        with open(format_file, 'r', encoding='utf-8') as f:
            fmt = list(DictReader(f))

        # table header
        hedaer_row = row
        view_rule = [r for r in fmt if r['visible'].lower() != 'false']
        for i, r in enumerate(view_rule, start=1):
            if r['view_name']:
                value = r['view_name']
            else:
                value = r['view_key']

            write_cell(ws, hedaer_row, i, value, int(rule_obj['header_width']))
            hedaer_row = hedaer_row + 1

    def gen_header(self) -> None:
        row = 1
        col = 1

        self.add_sheet_title('基本情報', '基本情報')

    def gen_global(self) -> None:
        global_obj = self.formatted_config_obj['global']

        for global_rule in self.format_rule_global:
            self.gen_category_sheet(global_rule, global_obj, mode='global')

    def gen_vdom(self) -> None:
        vdom_obj = self.formatted_config_obj['vdom']

        for vdom_name in vdom_obj:
            for vdom_rule in self.format_rule_vdom:
                self.gen_vdom_category_sheet(vdom_rule, vdom_obj[vdom_name], vdom_name)

    def export(self, formatted_config, export_dir='./') -> None:
        self.export_dir = export_dir
        self.rule_dir = Path(Path(__file__).parent.parent, 'data')
        self.global_rule_dir = Path(self.rule_dir, 'global')
        self.vdom_rule_dir = Path(self.rule_dir, 'vdom')
        self.formatted_config_obj = formatted_config
        self.gen_book_data()
        self.gen_header()
        self.gen_global()
        self.gen_vdom()
        self.save()


if __name__ == '__main__':
    import json
    json_path = r'C:\opt\work\phy\booya_config_parser\new_parser_vdom.json'
    with open(json_path, 'r', encoding='utf-8') as f:
        formatted_config = json.load(f)
    from config_parser.export.excel import ExportExcel
    excel = ExportExcel()
    excel.export(formatted_config)
