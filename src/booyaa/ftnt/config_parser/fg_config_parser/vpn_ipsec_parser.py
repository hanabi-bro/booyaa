import re
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class vpn_ipsec_parser():
    def __init__(self):
        self.vdom_conf = ''

        # phase1, phase2のフィールド読み込み
        import params.vpn_ipsec
        self.ipsec_categories = params.vpn_ipsec.categories
        # self.ipsec_categories = [
        #     'config vpn ipsec phase1-interface',
        #     'config vpn ipsec phase2-interface',
        # ]

        import params.config_vdom
        self.field_names = params.config_vdom.field_names

    def parse_params(self, vdom_conf):
        for vdom_key in [k for k in vdom_conf]:
            for target in self.ipsec_categories:
                list_params = []
                ### 未使用カテゴリのフォロー
                vdom_conf[vdom_key].setdefault(target, [])

                for l in vdom_conf[vdom_key][target]:
                    tmp_line = {}
                    for k, v in l.items():
                        tmp_line["id"] = [re.sub("edit ","" ,k).strip('"')]

                        for p in v:
                            tmp_line[p[1]] = p[2:]

                        self._set_default_value(tmp_line, target)                            

                        list_params.append(tmp_line)

                vdom_conf[vdom_key][target] = list_params
        
        self.vdom_conf = vdom_conf
        return vdom_conf

    def _set_default_value(self, tmp_line, target):
        tmp_line.setdefault('name', tmp_line["id"])
        filed_names = self.field_names[target]

        for fname in filed_names:
            tmp_line.setdefault(fname, filed_names[fname]['default_value'])
        
        return tmp_line

    def to_workbook(self, vdom_key, vdom_conf, wb):

        # sheet名はvdom_key ipsec vpn 追加
        vdom_name = re.sub("edit ","" ,vdom_key)
        sheet_name = f"{vdom_name} ipsec vpn"
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]


        # ヘッダー書き込み 
        ## 1行目
        row = 1
        col = 1
        ### Header
        ws.cell(row=row,column=1).value = sheet_name
        ws.cell(row=row,column=1).font = Font(bold=True, size='16')

        ## 1行開ける
        row += 2

        start_col = col
        start_row = row
        ph1_len = len(self.field_names[self.ipsec_categories[0]])

        for i, ph1_params in enumerate(vdom_conf[vdom_key][self.ipsec_categories[0]]):
            # phase1のフィールド書き出し
            row = start_row
            col = start_col
            for ph1_fname in self.field_names[self.ipsec_categories[0]]:
                row += 1
                ws.cell(row=row,column=col).value = ph1_fname
                ws.cell(row=row,column=col).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col)].width = 25

            # phase1のvalue書き出し
            row = start_row
            col += 1
            for ph1_fname in self.field_names[self.ipsec_categories[0]]:
                row += 1
                ws.column_dimensions[get_column_letter(col)].width = 30
                ws.cell(row=row,column=col).value = "\n".join(map(str, ph1_params[ph1_fname]))
                ws.cell(row=row,column=col).alignment = styles.Alignment(wrapText=True)

            # phase2のフィールド書き出し
            row = start_row
            col += 1
            for ph2_fname in self.field_names[self.ipsec_categories[1]]:
                row += 1
                ws.cell(row=row,column=col).value = ph2_fname
                ws.cell(row=row,column=col).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col)].width = 25

            # phase2のvalue書き出し
            ## phase1nameが同じPh2のパラメータを抜き出す（リレーションとPh2が複数の場合にも対応）
            ph2_conf = list(filter(lambda item : item['phase1name'] == ph1_params['name'], vdom_conf[vdom_key][self.ipsec_categories[1]]))
            for ph2_params in ph2_conf:
                row = start_row
                col += 1
                for ph2_fname in self.field_names[self.ipsec_categories[1]]:
                    row += 1
                    ws.column_dimensions[get_column_letter(col)].width = 30
                    ws.cell(row=row,column=col).value = "\n".join(map(str, ph2_params[ph2_fname]))
                    ws.cell(row=row,column=col).alignment = styles.Alignment(wrapText=True)
            
            # ph1 Tunnleが複数の場合下に追記していく
            start_row += ph1_len + 1

        return wb

if __name__ == '__main__':
    import pickle
    with open('./tmp/vdomconf.pkl', 'rb') as f:
        vdom_conf = pickle.load(f)

    ipsec = vpn_ipsec_parser()
    ipsec.parse_params(vdom_conf)

    wb = Workbook()
    wb.remove(wb.worksheets[0])

    wb = ipsec.to_workbook('edit root', ipsec.vdom_conf, wb)


    book_name = f"tmp/ipsec_debug.xlsx"
    wb.save(book_name)


