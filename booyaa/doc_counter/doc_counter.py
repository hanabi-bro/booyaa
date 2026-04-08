from pathlib import Path
from openpyxl import load_workbook
from zipfile import ZipFile
import xml.etree.ElementTree as ET
from pptx import Presentation

def get_file_list(root_dir, pattern, exclude):
    root_dir = Path(root_dir)

    files = []

    for p in root_dir.rglob(pattern):
        if any(
            p.name.startswith('~$') or
            parent.name in exclude or 
            parent.name.lower().startswith(tuple(exclude))
            for parent in p.parents
        ):
            continue
        files.append(p)
    
    return files

def count_page_excel(file):
    wb = load_workbook(file, read_only=True)

    # sheet_state == 'visible' だけを数える
    visible_sheets = [
        name for name in wb.sheetnames
        if wb[name].sheet_state == "visible"
    ]

    return len(visible_sheets)

def count_page_word(file):
    try:
        with ZipFile(file) as z:
            with z.open("docProps/app.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"d": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"}

                pages = root.find("d:Pages", ns)
                if pages is not None:
                    return int(pages.text)
                return 0
    except Exception as e:
        return 0

def count_page_powerpoint(file):
    try:
        prs = Presentation(file)
        return len(prs.slides)
    except Exception as e:
        return 0

def count_doc(root_dir, exclude=['old', '.old', '_old']):
    root_dir = Path(root_dir)
    # excel
    excel_sheet_count = 0
    excel_file_list = get_file_list(root_dir, '*.xlsx', exclude)
    for f in excel_file_list:
        excel_sheet_count += count_page_excel(f)

    # word
    word_page_count = 0
    word_file_list = get_file_list(root_dir, '*.docx', exclude)
    for f in word_file_list:
        word_page_count += count_page_word(f)

    # powrpoint
    powerpoint_page_count = 0
    powerpoint_file_list = get_file_list(root_dir, '*.pptx', exclude)
    for f in powerpoint_file_list:
        powerpoint_page_count += count_page_powerpoint(f)

    return {
        'excel_file_count': len(excel_file_list),
        'excel_sheet_count': excel_sheet_count,
        'word_file_count': len(word_file_list),
        'word_page_count': word_page_count,
        'powerpint_file_count': len(powerpoint_file_list),
        'powerpint_page_count': powerpoint_page_count,
    }

if __name__ == '__main__':
    from pprint import pprint
    import argparse    # 1. argparseをインポート

    parser = argparse.ArgumentParser(description='document num culc')
    parser.add_argument('dir', help='dir', default='.')

    args = parser.parse_args()

    let = count_doc(args.dir)
    pprint(let)