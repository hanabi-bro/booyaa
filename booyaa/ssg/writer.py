from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from os.path import join, abspath
from os import makedirs
from datetime import datetime


class ExportExcel():
    def __init__(self):
        self.__export_dir = abspath(join('.', 'export'))
        self.__config_orig = ''
        self.__config_parsed = {}
        self.__export_file_name = ''
        self.__export_file_path = ''

        self.__filed_names = {}

    def _parse_and_export(self):
        self._parse()

    def _parse(self):
        self._parse_basic(),
        self._parse_interface(),
        self._parse_address(),
        self._parse_service(),
        self._parse_service_group(),
        self._parse_vip(),
        self._parse_policy(),
        self._parse_route(),

    def timestamp(self):
        dt = datetime.now()
        return f'{dt:%Y%m%d_%H%M%S}'

    def _parse_basic(self):
        self.__filed_names['basic'] = {
            'hostname':     {'width': 25, 'default_value': []},
            'dns':          {'width': 25, 'default_value': []},
        }

        tmp_config = self.__config_orig['basic']
        for l in tmp_config:
            l['dns'] = '\n'.join(l['dns'])

        self.__config_parsed['basic'] = tmp_config

    def _parse_interface(self):
        self.__filed_names['interface'] = {
            'name':         {'width': 20, 'default_value': []},
            'address':      {'width': 25, 'default_value': []},
            'mode':         {'width': 10, 'default_value': []},
            'zone':         {'width': 10, 'default_value': []},
            'manageable':    {'width': 10, 'default_value': []},
            'gateway':      {'width': 20, 'default_value': []},
        }

        tmp_config = self.__config_orig['interface']
        for l in self.__config_orig['interface']:
            if 'ip' in l:
                l['address'] = l.pop('ip')

        self.__config_parsed['interface'] = tmp_config

    def _parse_route(self):
        self.__filed_names['route'] = {
            'dst':          {'width': 25, 'default_value': []},
            'gateway':      {'width': 25, 'default_value': []},
            'interface':    {'width': 20, 'default_value': []},
        }

        tmp_config = self.__config_orig['route']
        self.__config_parsed['route'] = tmp_config

    def _parse_policy(self):
        self.__filed_names['policy'] =  {
            'id':             {'width': 5,  'default_value': []},
            'status':         {'width': 10, 'default_value': []},
            'from_sz':        {'width': 15, 'default_value': []},
            'to_sz':          {'width': 15, 'default_value': []},
            'src-address':    {'width': 25, 'default_value': []},
            'dst-address':    {'width': 25, 'default_value': []},
            'service':        {'width': 20, 'default_value': []},
            'action':         {'width': 10, 'default_value': []},
            'comment':        {'width': 25, 'default_value': []},
            'log':            {'width': 10, 'default_value': []},
        }

        tmp_config = self.__config_orig['policy']
        for l in tmp_config:
            l['src-address'] = '\n'.join(l['src-address'])
            l['dst-address'] = '\n'.join(l['dst-address'])
            l['service'] = '\n'.join(l['service'])
            l['log'] = '\n'.join(l['log'])

        self.__config_parsed['policy'] = tmp_config

    def _parse_service(self):
        self.__filed_names['service'] = {
            'name':             {'width': 25, 'default_value': []},
            'protocol':         {'width':  5, 'default_value': []},
            'src-port':         {'width': 20, 'default_value': []},
            'dst-port':         {'width': 20, 'default_value': []},
            'comment':          {'width': 25, 'default_value': []},
        }

        tmp_config = self.__config_orig['service']
        for l in tmp_config:
            l['protocol'] = '\n'.join(l['protocol'])
            l['src-port'] = '\n'.join(l['src-port'])
            l['dst-port'] = '\n'.join(l['dst-port'])

        self.__config_parsed['service'] = tmp_config

    def _parse_service_group(self):
        self.__filed_names['service_group'] = {
            'name':             {'width': 25, 'default_value': []},
            'member':           {'width': 25, 'default_value': []},
            'comment':          {'width': 25, 'default_value': []},
        }

        # ['name', 'member', 'comment']

        tmp_config = self.__config_orig['service_group']

        for l in tmp_config:
            l['member'] = '\n'.join(l['member'])

        self.__config_parsed['service_group'] = tmp_config

    def _parse_address(self):
        self.__filed_names['address'] = {
            'name':             {'width': 25, 'default_value': []},
            'zone':             {'width': 10, 'default_value': []},
            'address':          {'width': 25, 'default_value': []},
            'comment':          {'width': 25, 'default_value': []},
        }

        tmp_config = self.__config_orig['address']
        self.__config_parsed['address'] = tmp_config

    def _parse_vip(self):
        self.__filed_names['vip'] = {
            'name':             {'width': 25, 'default_value': []},
            'vip':              {'width': 25, 'default_value': []},
            'mapped_address':   {'width': 25, 'default_value': []},
            'vr':               {'width': 10, 'default_value': []},
            'comment':          {'width': 25, 'default_value': []},
        }
        #['name', 'vip', 'mapped_address', 'vr', 'comments']

        tmp_config = self.__config_orig['vip']
        self.__config_parsed['vip'] = tmp_config

    def _mkdir(self):
        makedirs(self.export_dir, exist_ok=True)

    def _gen_export_file_name(self, file_name=None):
        if not file_name:
            file_name = f"{self.__config_orig['basic'][0]['hostname']}_config_{self.timestamp()}.xlsx"

        self.__export_file_name = file_name
        self.__export_file_path = abspath(join(self.__export_dir, file_name))

    @property
    def export_dir(self):
        return self.__export_dir

    @export_dir.setter
    def export_dir(self, export_dir):
        self.__export_dir = export_dir

    @property
    def export_file_name(self):
        return self.__export_file_name

    @export_file_name.setter
    def export_file_name(self, export_file_name):
        self.__export_file_name = export_file_name

    @property
    def config_orig(self):
        return self.__config_orig

    @config_orig.setter
    def config_orig(self, config_orig):
        self.__config_orig = config_orig

    @property
    def config_parsed(self):
        return self.__config_parsed

    def parse_to_excel_params(self, config):
        self.__config_orig = config
        self._parse()

    def export_excel_book(self):
        categories = [
            'basic',
            'interface',
            'address',
            'service',
            'service_group',
            'vip',
            'policy',
            'route',
        ]
        # workbook作成
        wb = Workbook()
        wb.remove(wb.worksheets[0])
        makedirs(self.export_dir, exist_ok=True)

        for category in categories:
            sheet_name = category
            wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            row = 1
            ### Header
            ws.cell(row=row,column=1).value = sheet_name
            ws.cell(row=row,column=1).font = Font(bold=True, size='16')

            # 1行空ける
            row = row + 2

            # ヘッダ行
            for col, field_name in enumerate(self.__filed_names[category], start=1):
                    ws.cell(row=row,column=col).value = field_name
                    ws.cell(row=row,column=col).font = Font(bold=True)

                    ws.column_dimensions[get_column_letter(col)].width = self.__filed_names[category][field_name]['width']

            # params
            for r, param in enumerate(self.__config_parsed[category], start=row+1):
                for col, field_name in enumerate(self.__filed_names[category], start=1):
                    ws.cell(row=r,column=col).value = param[field_name]
                    ws.cell(row=r,column=col).alignment = styles.Alignment(wrapText=True)

        # workbook保存
        self._gen_export_file_name()
        wb.save(self.__export_file_path)
